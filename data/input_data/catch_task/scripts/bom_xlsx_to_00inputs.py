#!/usr/bin/env python3
"""Turn a supporting-equipment xlsx into thermal 00_inputs.

Simple policy:
- xlsx is authoritative for mass and power.
- thermal database is used as a shape/material/CAD reference.
- match CATCH records by name first.
- if not found, match by inferred kind/subsystem and closest dimensions.
- copy a template 00_inputs directory and replace only:
  real_bom.items, geom.components, layout_topology.placements.
"""

from __future__ import annotations

import argparse
import copy
import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_DB = REPO_ROOT / "backend/workflow_agents/thermal_skills/config-editor/references/热仿真数据库.json"
DEFAULT_TEMPLATE = REPO_ROOT / "data/input_data/thermal_catch/00_inputs"

KIND_WORDS = [
    ("星箭分离", "mechanism", "separation_device"),
    ("行程开关", "mechanism", "limit_switch"),
    ("伸展", "mechanism", "deployment"),
    ("展开", "mechanism", "deployment"),
    ("反作用", "adcs", "reaction_wheel"),
    ("飞轮", "adcs", "reaction_wheel"),
    ("星敏", "adcs", "star_tracker"),
    ("太阳敏感器", "adcs", "sun_sensor"),
    ("磁强计", "adcs", "magnetometer"),
    ("磁力矩器", "adcs", "magnetorquer"),
    ("磁棒", "adcs", "magnetorquer"),
    ("陀螺", "adcs", "gyro"),
    ("电推", "propulsion", "thruster"),
    ("微推", "propulsion", "thruster"),
    ("推力器", "propulsion", "thruster"),
    ("电池", "power", "battery"),
    ("太阳电池阵", "power", "solar_array"),
    ("帆板", "power", "solar_array"),
    ("综合电子", "avionics", "electronics_box"),
    ("星务", "avionics", "onboard_computer"),
    ("测控数传", "communication", "ttc_box"),
    ("短报文", "communication", "communication_box"),
    ("GNSS", "communication", "gnss"),
    ("天线", "communication", "antenna"),
    ("微波开关", "communication", "microwave_switch"),
    ("探测器", "payload", "detector"),
    ("相机", "payload", "camera"),
    ("光学", "payload", "optical_payload"),
    ("载荷", "payload", "payload"),
    ("热控", "thermal", "thermal_control"),
    ("加热", "thermal", "heater"),
]

SUBSYSTEM_KIND = {
    "结构与机构分系统": ("mechanism", "mechanism"),
    "机构分系统": ("mechanism", "mechanism"),
    "姿轨控分系统": ("adcs", "adcs"),
    "推进分系统": ("propulsion", "propulsion"),
    "电源与总体电路分系统": ("power", "power"),
    "电源分系统": ("power", "power"),
    "综合电子分系统": ("avionics", "avionics"),
    "测控 / 数传一体机分系统": ("communication", "communication"),
    "载荷分系统": ("payload", "payload"),
    "热控分系统": ("thermal", "thermal"),
}


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def norm(text: Any) -> str:
    text = "" if text is None else str(text)
    text = re.sub(r"^CATCH-P\d+\s*", "", text.strip(), flags=re.I)
    return re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff]+", "", text).lower()


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if isinstance(value, float) and math.isnan(value) else float(value)
    text = str(value).replace(",", "")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    value = float(match.group(0))
    if re.search(r"\bmw\b", text, re.I):
        return value / 1000.0
    return value


def power(value: Any) -> float | None:
    if isinstance(value, str) and "=" in value:
        values = [float(x) for x in re.findall(r"[-+]?\d+(?:\.\d+)?", value)]
        return values[-1] if values else None
    return num(value)


def dims(value: Any) -> list[float] | None:
    if value is None:
        return None
    values = [float(x) for x in re.findall(r"[-+]?\d+(?:\.\d+)?", str(value).replace(",", ""))]
    return values[:3] if len(values) >= 3 and all(v > 0 for v in values[:3]) else None


def record_dims(record: dict[str, Any]) -> list[float] | None:
    values = [num(record.get(k)) for k in ("长 mm", "宽 mm", "高 mm")]
    if all(v and v > 0 for v in values):
        return [float(v) for v in values]
    values = [num(record.get(k)) for k in ("STEP长", "STEP宽", "STEP高")]
    if all(v and v > 0 for v in values):
        return [float(v) for v in values]
    return dims(record.get("尺寸"))


def infer_kind(name: str, subsystem: str | None = None, db_kind: str | None = None) -> tuple[str, str]:
    text = f"{name} {db_kind or ''}"
    for word, category, subtype in KIND_WORDS:
        if word.lower() in text.lower():
            return category, subtype
    if subsystem in SUBSYSTEM_KIND:
        return SUBSYSTEM_KIND[subsystem]
    return "payload", "payload"


def load_table(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = {str(c.value).strip(): i + 1 for i, c in enumerate(ws[1]) if c.value}

    def cell(row: int, name: str) -> Any:
        col = headers.get(name)
        return ws.cell(row, col).value if col else None

    rows: list[dict[str, Any]] = []
    subsystem: str | None = None
    for r in range(2, ws.max_row + 1):
        name = str(cell(r, "产品名称") or "").strip()
        if not name:
            continue
        mass = num(cell(r, "重量（Kg）"))
        size = dims(cell(r, "包络尺寸（mm）"))
        avg_power = power(cell(r, "稳态功耗（W）"))
        peak_power = power(cell(r, "峰值功耗（W）"))
        if name in SUBSYSTEM_KIND:
            subsystem = name
            continue
        if name == "整星质量":
            continue
        if mass is None and size is None and avg_power is None and peak_power is None:
            continue
        category, subtype = infer_kind(name, subsystem)
        rows.append(
            {
                "row": r,
                "name": name,
                "mass_kg": mass if mass is not None else 0.0,
                "dims_mm": size,
                "power_W": avg_power if avg_power is not None else 0.0,
                "peak_power_W": peak_power,
                "subsystem": subsystem,
                "category": category,
                "subtype": subtype,
            }
        )
    return rows


def load_db(path: Path) -> list[dict[str, Any]]:
    db = read_json(path)
    out: list[dict[str, Any]] = []
    for sheet in db.get("sheets", []):
        for i, record in enumerate(sheet.get("records", [])):
            if i == 0 and record.get("器件型号") == "model":
                continue
            names = [record.get(k) for k in ("器件型号", "器件名称", "器件名称(中文)", "器件ID")]
            names = [str(v).strip() for v in names if v]
            if not names:
                continue
            category, subtype = infer_kind(
                " ".join(names),
                str(record.get("所属分系统") or ""),
                str(record.get("器件种类") or ""),
            )
            source = str(record.get("器件来源") or "")
            out.append(
                {
                    "record": record,
                    "sheet": sheet.get("name"),
                    "index": i,
                    "names": names,
                    "norm_names": [norm(v) for v in names],
                    "dims_mm": record_dims(record),
                    "is_catch": source.upper() == "CATCH" or source.lower().startswith("catch"),
                    "category": category,
                    "subtype": subtype,
                }
            )
    return out


def dim_score(a: list[float] | None, b: list[float] | None) -> float:
    if not a or not b:
        return 99.0
    aa, bb = sorted(a), sorted(b)
    return sum(abs(x - y) / max(x, y, 1.0) for x, y in zip(aa, bb))


def match_row(row: dict[str, Any], db_rows: list[dict[str, Any]]) -> tuple[dict[str, Any], str, float]:
    row_name = norm(row["name"])
    catch_hits = []
    for cand in db_rows:
        if cand["is_catch"] and any(row_name in n or n in row_name for n in cand["norm_names"] if n):
            catch_hits.append((dim_score(row["dims_mm"], cand["dims_mm"]), cand))
    if catch_hits:
        score, cand = min(catch_hits, key=lambda x: x[0])
        return cand, "catch_name", score

    pool = [
        c
        for c in db_rows
        if c["subtype"] == row["subtype"] or c["category"] == row["category"]
    ] or db_rows
    scored = [(dim_score(row["dims_mm"], c["dims_mm"]), c) for c in pool]
    score, cand = min(scored, key=lambda x: x[0])
    return cand, "similar_kind_size", score


def template_maps(real_bom: dict[str, Any], geom: dict[str, Any]) -> dict[str, dict[str, Any]]:
    geom_by_component = {
        comp.get("component_id"): comp for comp in geom.get("components", {}).values() if comp.get("component_id")
    }
    by_name = {}
    for item in real_bom.get("items", []):
        comp_id = item.get("component_id")
        comp = geom_by_component.get(comp_id)
        if not comp:
            continue
        for value in (item.get("semantic_name"), (item.get("source_ref") or {}).get("display_name")):
            if value:
                by_name[norm(value)] = {"item": item, "geom": comp}
    return by_name


def axes_from_face(face_id: str) -> tuple[str, int, int, int]:
    local = face_id.split(".local_", 1)[-1] if ".local_" in face_id else "zmax"
    axis = "xyz".find(local[0])
    axis = axis if axis >= 0 else 2
    sign = 1 if local.endswith("max") else -1
    axes = [0, 1, 2]
    axes.remove(axis)
    return local, axis, sign, axes[0], axes[1]


def first_face(layout: dict[str, Any]) -> dict[str, Any]:
    faces = layout.get("install_faces") or []
    outer = [f for f in faces if f.get("side") == "outer"]
    return (outer or faces)[0]


def synthetic_box(face: dict[str, Any], size: list[float], index: int) -> dict[str, list[float]]:
    axis = int(face.get("plane_axis", 2))
    sign = int(face.get("normal_sign", 1))
    axes = [0, 1, 2]
    axes.remove(axis)
    u, v = axes
    col, row = index % 5, index // 5
    center = [0.0, 0.0, 0.0]
    center[u] = -120 + col * 60
    center[v] = -120 + (row % 5) * 60
    plane = float(face.get("plane_value", 0.0))
    mn = [center[i] - size[i] / 2 for i in range(3)]
    mx = [center[i] + size[i] / 2 for i in range(3)]
    if sign >= 0:
        mn[axis], mx[axis] = plane + 1.0, plane + 1.0 + size[axis]
    else:
        mn[axis], mx[axis] = plane - 1.0 - size[axis], plane - 1.0
    return {"min": mn, "max": mx}


def make_outputs(xlsx: Path, db_path: Path, template_dir: Path, output_dir: Path) -> dict[str, Any]:
    rows = load_table(xlsx)
    db_rows = load_db(db_path)
    real_bom = read_json(template_dir / "real_bom.json")
    geom = read_json(template_dir / "geom.json")
    layout = read_json(template_dir / "layout_topology.json")
    templates = template_maps(real_bom, geom)

    real_bom = copy.deepcopy(real_bom)
    geom = copy.deepcopy(geom)
    layout = copy.deepcopy(layout)
    real_bom["items"] = []
    geom["components"] = {}
    layout["placements"] = []

    report = []
    default_face = first_face(layout)

    for i, row in enumerate(rows, start=1):
        match, mode, score = match_row(row, db_rows)
        record = match["record"]
        size = row["dims_mm"] or match["dims_mm"] or [50.0, 50.0, 50.0]
        comp_id = f"P{i:03d}"
        geom_id = f"G{i:03d}"
        thermal_id = f"T{i:03d}"

        tpl = templates.get(norm(row["name"]))
        if tpl:
            comp = copy.deepcopy(tpl["geom"])
            face_id = str(comp.get("component_mount_face_id", f"{comp_id}.local_zmax"))
            old_id = str(comp.get("component_id", ""))
            if old_id and face_id.startswith(old_id + "."):
                face_id = comp_id + face_id[len(old_id) :]
            bbox_min = list((comp.get("bbox") or {}).get("min") or comp.get("position") or [0, 0, 0])
            bbox = {"min": bbox_min, "max": [bbox_min[j] + size[j] for j in range(3)]}
            mount_face_id = comp.get("mount_face_id")
            kind = comp.get("kind", "external")
        else:
            face_id = f"{comp_id}.local_zmax"
            bbox = synthetic_box(default_face, size, i - 1)
            mount_face_id = default_face["id"]
            kind = "external"
            comp = {}

        geom_key = f"{geom_id}_{comp_id}"
        comp.update(
            {
                "id": geom_key,
                "component_id": comp_id,
                "semantic_name": row["name"],
                "kind": kind,
                "category": row["category"],
                "component_subtype": row["subtype"],
                "dims": size,
                "mass": row["mass_kg"],
                "power": row["power_W"],
                "shape": record.get("外形") or comp.get("shape") or "box",
                "bbox": bbox,
                "position": bbox["min"],
                "mount_face_id": mount_face_id,
                "component_mount_face_id": face_id,
                "thermal_surface": comp.get("thermal_surface") or {"absorptivity": 0.3, "emissivity": num(record.get("辐射率")) or 0.8},
                "thermal_interface": comp.get("thermal_interface") or {"contact_resistance": num(record.get("接触热阻K/W")) or 0.001},
            }
        )
        geom["components"][geom_key] = comp

        local, axis, sign, u_axis, v_axis = axes_from_face(face_id)
        real_bom["items"].append(
            {
                "component_id": comp_id,
                "semantic_name": row["name"],
                "kind": kind,
                "category": row["category"],
                "size_mm": size,
                "mass_kg": row["mass_kg"],
                "power_W": row["power_W"],
                "peak_power_W": row["peak_power_W"],
                "material_id": "aluminum_6061",
                "mounting": {
                    "default_component_mount_face_id": face_id,
                    "mount_faces": [
                        {
                            "component_mount_face_id": face_id,
                            "local_face": local,
                            "normal_axis": axis,
                            "normal_sign": sign,
                            "u_axis": u_axis,
                            "v_axis": v_axis,
                        }
                    ],
                },
                "quantity": 1,
                "source_ref": {
                    "supporting_table_row": row["row"],
                    "matched_sheet": match["sheet"],
                    "matched_row_index": match["index"],
                    "matched_model": record.get("器件型号"),
                    "matched_name": record.get("器件名称"),
                    "matched_name_cn": record.get("器件名称(中文)"),
                    "matched_source": record.get("器件来源"),
                    "cad_path": record.get("CAD路径"),
                    "cad_rotated_path": record.get("CAD_rotated_path") or record.get("Rotated CAD Path"),
                    "cad_major_path": record.get("CAD_MAJOR_PATH"),
                },
            }
        )

        layout["placements"].append(
            {
                "component_id": comp_id,
                "semantic_name": row["name"],
                "kind": kind,
                "cabin_id": None,
                "component_mount_face_id": face_id,
                "mount_face_id": mount_face_id,
                "alignment": {"normal_alignment": "opposite", "in_plane_rotation_deg": 0.0},
                "geometry_id": geom_id,
                "thermal_id": thermal_id,
                "category": row["category"],
            }
        )
        report.append(
            {
                "row": row["row"],
                "name": row["name"],
                "component_id": comp_id,
                "match_mode": mode,
                "match_score": round(score, 6),
                "matched_model": record.get("器件型号"),
                "matched_name_cn": record.get("器件名称(中文)"),
                "matched_source": record.get("器件来源"),
                "mass_kg": row["mass_kg"],
                "power_W": row["power_W"],
                "peak_power_W": row["peak_power_W"],
                "size_mm": size,
            }
        )

    real_bom["bom_id"] = f"{xlsx.stem}_generated_bom"
    real_bom["source"] = {"type": "supporting_table", "xlsx": str(xlsx), "database": str(db_path)}
    geom.setdefault("meta", {})["source_supporting_table"] = str(xlsx)
    layout["layout_id"] = f"{xlsx.stem}_generated_layout"
    layout["source_design_id"] = "supporting_table"

    write_json(output_dir / "real_bom.json", real_bom)
    write_json(output_dir / "geom.json", geom)
    write_json(output_dir / "layout_topology.json", layout)
    write_json(output_dir / "match_report.json", {"component_count": len(rows), "matches": report})
    return {"output_dir": str(output_dir), "component_count": len(rows)}


def generate_00inputs_from_supporting_table(
    xlsx_path: str | Path,
    output_dir: str | Path,
    db_path: str | Path = DEFAULT_DB,
    template_dir: str | Path = DEFAULT_TEMPLATE,
) -> dict[str, Any]:
    return make_outputs(Path(xlsx_path), Path(db_path), Path(template_dir), Path(output_dir))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--template-dir", type=Path, default=DEFAULT_TEMPLATE)
    args = parser.parse_args()
    result = generate_00inputs_from_supporting_table(args.xlsx_path, args.output_dir, args.db, args.template_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
