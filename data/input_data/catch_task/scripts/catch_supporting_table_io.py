#!/usr/bin/env python3
"""Read/write CATCH supporting-table xlsx and refresh thermal 00_inputs."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from bom_xlsx_to_00inputs import generate_00inputs_from_supporting_table


HEADERS = ["产品名称", "重量（Kg）", "包络尺寸（mm）", "稳态功耗（W）", "峰值功耗（W）", "工作温度（℃）", "配套单位"]


def read_table(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, Any]] = []
    for row_index, values in enumerate(worksheet.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True), start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append({
            "id": f"r{row_index}",
            "row": row_index,
            "产品名称": values[0],
            "重量（Kg）": values[1],
            "包络尺寸（mm）": values[2],
            "稳态功耗（W）": values[3],
            "峰值功耗（W）": values[4],
            "工作温度（℃）": values[5],
            "配套单位": values[6],
        })
    return {"headers": HEADERS, "rows": rows, "source_path": str(path)}


def write_table(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CATCH整星配套表"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append([row.get(header) for header in HEADERS])
    for column, width in zip("ABCDEFG", [32, 14, 22, 14, 14, 18, 16]):
        worksheet.column_dimensions[column].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    read_cmd = sub.add_parser("read")
    read_cmd.add_argument("xlsx", type=Path)
    write_cmd = sub.add_parser("write-refresh")
    write_cmd.add_argument("xlsx", type=Path)
    write_cmd.add_argument("output_dir", type=Path)
    write_cmd.add_argument("--rows-json", required=True)
    args = parser.parse_args()

    if args.command == "read":
        print(json.dumps(read_table(args.xlsx), ensure_ascii=False))
        return

    rows = json.loads(args.rows_json)
    if not isinstance(rows, list):
        raise SystemExit("rows-json must be an array")
    write_table(args.xlsx, rows)
    result = generate_00inputs_from_supporting_table(args.xlsx, args.output_dir)
    print(json.dumps({"table": read_table(args.xlsx), "generation": result}, ensure_ascii=False))


if __name__ == "__main__":
    main()
