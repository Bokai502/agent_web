from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def normalize_excel_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            number = float(text)
            if number.is_integer():
                return str(int(number))
        except Exception:
            return text
    return text


def last_non_empty_column(rows: list[list[str]]) -> int:
    width = max((len(row) for row in rows), default=0)
    while width > 0:
        index = width - 1
        if any(index < len(row) and row[index] for row in rows):
            return width
        width -= 1
    return 0


def read_xlsx_table(path: Path, max_rows: int = 120, max_cols: int = 16) -> dict[str, Any]:
    if not path.exists():
        return {"path": str(path), "exists": False, "headers": [], "rows": []}
    if load_workbook is None:
        return {"path": str(path), "exists": True, "error": "openpyxl is not installed", "headers": [], "rows": []}
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {
            "path": str(path),
            "exists": True,
            "error": f"{type(exc).__name__}: {exc}",
            "headers": [],
            "rows": [],
        }
    try:
        worksheet = next((sheet for sheet in workbook.worksheets if sheet.max_row and sheet.max_column), workbook.active)
        raw_rows: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            cells = [normalize_excel_cell(value) for value in row[:max_cols]]
            if any(cells):
                raw_rows.append(cells)
            if len(raw_rows) >= max_rows + 1:
                break
        if not raw_rows:
            return {"path": str(path), "exists": True, "sheet": worksheet.title, "headers": [], "rows": []}
        max_width = last_non_empty_column(raw_rows)
        trimmed = [(row + [""] * max_width)[:max_width] for row in raw_rows]
        return {
            "path": str(path),
            "exists": True,
            "sheet": worksheet.title,
            "headers": trimmed[0],
            "rows": trimmed[1:],
        }
    finally:
        workbook.close()


def build_thermal_control_table(catch_table: dict[str, Any]) -> dict[str, Any]:
    headers = catch_table.get("headers")
    rows = catch_table.get("rows")
    if not isinstance(headers, list) or not isinstance(rows, list):
        return {"headers": ["序号", "仪器设备名称", "热控指标"], "rows": []}
    try:
        name_index = headers.index("产品名称")
    except ValueError:
        name_index = 0
    try:
        temperature_index = headers.index("工作温度（℃）")
    except ValueError:
        temperature_index = -1
    output_rows: list[list[str]] = []
    for row in rows:
        if not isinstance(row, list):
            continue
        name = row[name_index] if name_index < len(row) else ""
        if not str(name).strip():
            continue
        temperature = row[temperature_index] if temperature_index >= 0 and temperature_index < len(row) else ""
        output_rows.append([str(len(output_rows) + 1), str(name), str(temperature)])
    return {
        "headers": ["序号", "仪器设备名称", "热控指标"],
        "rows": output_rows,
    }
